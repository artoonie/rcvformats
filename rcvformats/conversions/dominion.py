"""
Reads an Dominion XLSX results file, writes to the standard format
"""
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.read_only import EmptyCell, ReadOnlyCell

from rcvformats.conversions.base import GenericGuessAtTransferConverter


class DominionConverter(GenericGuessAtTransferConverter):
    """
    Parses the dominion file format as exemplified in /testdata/inputs/dominion
    These are .xlsx files
    """

    # Define constants
    DATE_CELL = 'A9'

    class RoundInfo:  # pylint: disable=too-few-public-methods
        """
        Data for parsing a round:
        Because columns are hapharzadly merged, there is no straightforward mapping from
        round number to the corresponding column. This little struct keeps track of it for us.
        """
        round_num = None
        column = None

    class RowConstants():
        """
        Data for the row number that various items are on
        """
        # Define constants
        # the -12 is because New Mexico files have 12 rows of headers,
        # so the first number is the actual row number, then we subtract num_header_rows,
        # just for readability.
        SEAT_TITLE_NUM_ROWS_AFTER_HEADER = 12 - 12
        ROUND_LABELS_NUM_ROWS_AFTER_HEADER = 32 - 12
        FIRST_CANDIDATE_NUM_ROWS_AFTER_HEADER = 34 - 12
        ROW_AFTER_INACTIVE_FOR_THRESHOLD = 1

        def __init__(self):
            # The row that includes the seat title in column 1
            self.seat_title = None

            # The row that includes the round labels in the primary table
            self.round_label = None

            # The row that includes the candidate names in column 1
            self.first_candidate = None

            # Row that has the threshold at each round, if one is provided at all
            self.maybe_threshold = None

            # Row that has the number of inactive ("non transferrable") ballots at each round
            self.inactive_ballots = None

        def find_rows_before_summary_table(self, sheet):
            """
            Fills in values for rows before the summary table
            """
            offset = self._count_num_header_rows(sheet)
            self.round_label = self.ROUND_LABELS_NUM_ROWS_AFTER_HEADER + offset
            self.first_candidate = self.FIRST_CANDIDATE_NUM_ROWS_AFTER_HEADER + offset
            self.seat_title = self.SEAT_TITLE_NUM_ROWS_AFTER_HEADER + offset

        def find_rows_after_summary_table(self, sheet, num_candidates):
            """
            Fills in values for rows after the summary table,
            which requires needing to know the number of candidates
            """
            self.inactive_ballots = self._find_row_of_inactive_ballots(sheet, num_candidates)
            self.maybe_threshold = self._try_to_find_row_of_threshold(sheet, self.inactive_ballots)

        @classmethod
        def _count_num_header_rows(cls, sheet):
            """
            The number of header rows can vary.
            Looks for the first left-aligned row, which is row 12 or 13 in all data
            we've seen.
            """
            min_header_rows = 3
            max_header_rows = 50
            for row in range(min_header_rows, max_header_rows):
                # Empty rows are merged rows - ignore them
                if sheet.cell(row, 1).value is None:
                    continue

                # Center-aligned rows are header rows - keep going
                if sheet.cell(row, 1).alignment.horizontal == 'center':
                    continue

                return row
            raise Exception("Could not find the end of the headers")

        def _find_row_of_inactive_ballots(self, sheet, num_candidates):
            """
            Returns row number labeled "Non-Transferrable Total"
            """
            row = self.first_candidate
            row += num_candidates

            min_num_rows = 1
            max_num_rows = 6
            for row in range(row + min_num_rows, row + max_num_rows):
                if sheet.cell(row, 1).value == "Non Transferable Total":
                    return row
            raise Exception("Could not find the end of the non-transferable rows")

        def _try_to_find_row_of_threshold(self, sheet, inactive_row):
            """
            Returns row number labeled "Threshold"
            """
            row = inactive_row + self.ROW_AFTER_INACTIVE_FOR_THRESHOLD
            if sheet.cell(row, 1).value == "Threshold":
                return row
            return None

    def __init__(self):
        super().__init__()

        # These fields will be filled out as data is loaded

        # The loaded spreadsheet
        self.sheet = None

        # Row constants
        self.row_constants = self.RowConstants()

        # names of each candidate
        self.candidates = None

        # list of RoundInfo, one per round
        self.data_per_round = None

    def _convert_file_object_to_ut(self, file_object):
        workbook = load_workbook(file_object)  # note: somehow, readonly is 2x slower
        self.sheet = workbook.active

        self.row_constants.find_rows_before_summary_table(self.sheet)
        self.candidates = self._parse_candidates()
        self.row_constants.find_rows_after_summary_table(self.sheet, len(self.candidates))

        self.data_per_round = self._parse_rounds()

        config = self._parse_config()
        results = self._get_vote_counts_per_candidate()
        urcvt_data = {'config': config, 'results': results}

        workbook.close()

        self._postprocess_remove_last_round_elimination(urcvt_data)
        self._postprocess_set_threshold(urcvt_data)
        return urcvt_data

    def _parse_config(self):
        """
        Returns the URCV config format
        """
        config = {}

        # First, try to parse the date
        date_with_time = self.sheet[self.DATE_CELL].value
        try:
            config['date'] = str(date_with_time.date())
        except AttributeError:
            # No date - as per the SF final short report
            pass

        # Then, grab the required title
        seat = self.sheet.cell(self.row_constants.seat_title, 1).value
        config['contest'] = seat
        config['office'] = seat

        return config

    def _parse_rounds(self):
        """
        Rounds are curious - they are separated by a varying number of columns,
        usually 3 except for the first few, where the initial columns have been
        merged seemingly randomly (but in actually: twice in round 1, once in round 2)

        Therefore, this returns a pair of (Round, Column #) to get the column for
        the number of votes in each Round
        """
        col = 3
        data_per_round = []
        max_cols = 1500
        while True:
            cell = self.sheet.cell(self.row_constants.round_label, col)
            value = cell.value

            col += 1
            if col >= max_cols:
                raise Exception("This document is not in the correct format..."
                                "or there are more than 500 rounds")

            # Is this a merged cell? If so, ignore it.
            if isinstance(cell, MergedCell):
                continue

            # Have we reached the end of the table?
            if isinstance(cell, EmptyCell):
                break

            if value is None:
                # If we open this as read-only, then None represents a
                # merged cell and we should keep reading
                if isinstance(cell, ReadOnlyCell):
                    continue

                # Otherwise, it represents the end of the table
                break

            num_rounds = len(data_per_round)

            # Are we still in the same round?
            if value == 'Round ' + str(num_rounds):
                continue

            # If not, we must be in the next round. Increment and carry on.
            assert value == 'Round ' + str(num_rounds + 1)
            roundinfo = self.RoundInfo()
            roundinfo.round_num = num_rounds
            roundinfo.column = col - 1
            data_per_round.append(roundinfo)

        assert len(data_per_round) >= 1
        return data_per_round

    def _parse_candidates(self):
        """
        Grabs the list of candidate names from the summary table
        """
        end_of_candidates_marker = 'Continuing Ballots Total'
        max_num_rows = 500
        candidate_names = []
        row = self.row_constants.first_candidate
        while True:
            candidate_name = self.sheet['A' + str(row)].value
            if candidate_name == end_of_candidates_marker:
                break
            if row == max_num_rows:
                raise Exception("This document is not in the correct format..."
                                "or there are more than 500 candidates")
            cell = self.sheet['A' + str(row)]
            name = cell.value
            candidate_names.append(name)
            row += 1
        assert len(candidate_names) >= 1
        return candidate_names

    @classmethod
    def _is_eliminated_color(cls, color):
        """
        Is this cell colored red for elimination?
        Note: only call this if you've checked that the cell is filled solid.
        It may be the case that the cell is red, but not filled, in which case
        it appears white and is not eliminated.
        """
        return color == 'FFFFABAB'

    @classmethod
    def _is_elected_color(cls, color):
        """
        Is this cell colored greed, noting it is elected?
        Note: only call this if you've checked that the cell is filled solid.
        It may be the case that the cell is red, but not filled, in which case
        it appears white and is not eliminated.
        """
        return color == 'FF89CC89'

    def _parse_tally_for_round_at_column(self, col, eliminated_names):
        """ Creates a 'tally' and 'tallyResults' struct for the given round """
        starting_candidate_row = self.row_constants.first_candidate
        tally = {}
        tally_results = []
        for i, name in enumerate(self.candidates):
            if name in eliminated_names:
                continue

            row = starting_candidate_row + i
            cell = self.sheet.cell(row, col)
            vote_count = cell.value
            tally[name] = vote_count

            fill_type = cell.fill.fill_type
            if fill_type is not None:
                bg_color = cell.fill.bgColor.rgb
                if self._is_eliminated_color(bg_color):
                    eliminated_names.add(name)
                    tally_results.append({'eliminated': name})
                elif self._is_elected_color(bg_color):
                    tally_results.append({'elected': name})
                else:
                    # Just make sure we have no rogue colors
                    assert bg_color == '00000000'

        # Add inactive ballots
        tally['Inactive Ballots'] = self.sheet.cell(self.row_constants.inactive_ballots, col).value

        return tally, tally_results, eliminated_names

    def _get_vote_counts_per_candidate(self):
        results = []
        eliminated_names = set()
        for round_info in self.data_per_round:
            col = round_info.column
            tally, tally_results, eliminated_names = \
                self._parse_tally_for_round_at_column(col, eliminated_names)
            results.append({
                'round': round_info.round_num + 1,
                'tally': tally,
                'tallyResults': tally_results})
        return results

    @classmethod
    def _postprocess_remove_last_round_elimination(cls, data):
        """
        When there are two candidates left, Dominion marks the loser among them as
        "eliminated", whereas the URCVT format does not.
        Updates data to remove any last-round eliminations
        """
        last_round_tally_results = data['results'][-1]['tallyResults']
        last_round_tally_results = [t for t in last_round_tally_results if 'eliminated' not in t]
        data['results'][-1]['tallyResults'] = last_round_tally_results

    def _postprocess_set_threshold(self, data):
        """
        Set the threshold based on (last round active votes) / (num winners + 1)
        Which is also just listed on the table of per-round info
        """
        last_round_col = self.data_per_round[-1].column
        maybe_threshold_row = self.row_constants.maybe_threshold
        if maybe_threshold_row is not None:
            threshold = self.sheet.cell(maybe_threshold_row, last_round_col).value
            data['config']['threshold'] = threshold
